from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options

from config.log_config import logger
import time
from bin.main import Method
from bin.login import Login
from new_selenium.test_weixiu.weixiu_config import *
from config.config import path_dir
import os
from openpyxl import load_workbook


class Tech:    
    def log_file_out(self,msg):
        fo = open(r'{}/usecase.txt'.format(path_dir), mode='a', encoding='utf-8')
        fo.write(msg + '\r\n')
        fo.close()
        
    def tech_fault_analysis(self,url):

        # option = webdriver.ChromeOptions()
        # option.add_argument("headless")
        # driver = webdriver.Chrome(chrome_options=option,
        #                           executable_path=r'C:\Users\a\PycharmProjects\untitled\new_selenium\apps\chromedriver.exe')
        #
        # options = webdriver.ChromeOptions()
        # prefs = {'profile.default_content_settings.popups': 0,
        #          'download.default_directory': '{}/tech_fault_number'.format(path_dir),
        #          "profile.default_content_setting_values.automatic_downloads": 1}
        # options.add_experimental_option('prefs', prefs)
        # driver = webdriver.Chrome(chrome_options=options,
        #                           executable_path=r'{}/apps/chromedriver.exe'.format(path_dir))
        chrome_options = Options()

        prefs = {'profile.default_content_settings.popups': 0,
                 'download.default_directory': '{}/tech_fault_number'.format(path_dir),
                 "profile.default_content_setting_values.automatic_downloads": 1,
                 "download.prompt_for_download": False, }
        chrome_options.add_experimental_option("prefs", prefs)

        # 无头模式（就是不打开浏览器）
        chrome_options.add_argument("--headless")
        driver = webdriver.Chrome(chrome_options=chrome_options)

        # 老外的样例，我先照着写，有没有大神可以解释一下的。下面的downloadpath要改成和上面一样
        driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
        params = {'cmd': 'Page.setDownloadBehavior', 'params': {'behavior': 'allow', 'downloadPath': "E:\download"}}
        command_result = driver.execute("send_command", params)
        Login().login(url,'test','1234',driver)

        self.log_file_out('---修程修制与技术更改故障单---')
        for i in contents:
            try:
                Method(driver).contains_xpath('click',i)
                self.log_file_out('点击'+i+'成功')
            except Exception as e:
                logger.debug(e)
                self.log_file_out('点击' + i + '失败')

        time.sleep(2)

        try:
            Method(driver).switch_out()
            Method(driver).switch_iframe(
                driver.find_element_by_xpath("//iframe[contains(@src,'/darams/a/mould/repairProcedure')]"))

            time.sleep(2)
        except NoSuchElementException as e:
            self.log_file_out('点击新建按钮失败,未找到立即计算按钮的xpath')

        try:
            driver.find_element_by_xpath("//a[contains(text(),\'{}\')]/../../td[7]/a[1]".format('导出故障单测试1')).click()
            self.log_file_out('导出修程修制故障单成功')
        except:
            self.log_file_out('导出修程修制故障单失败')

        time.sleep(25)
        try:
            dir_list = os.listdir(r'{}/tech_fault_number'.format(path_dir))
            dir_list = sorted(dir_list, key=lambda x: os.path.getctime(os.path.join(r'{}/tech_fault_number'.format(path_dir), x)))
            os.rename(r'{}/tech_fault_number/{}'.format(path_dir, dir_list[-1]),
                      r'{}/tech_fault_number/{}'.format(path_dir, '修程修制故障单.xlsx'))
        except:
            self.log_file_out('修程修制故障单重命名失败')

        time.sleep(3)
        try:
            wb = load_workbook(r'{}/tech_fault_number/修程修制故障单.xlsx'.format(path_dir))
            sheet = wb.active
            repair = ''
            repair_L = []
            for i in range(0, len(list(sheet.columns))):
                for column in list(sheet.columns)[i]:
                    if column.value == '累计里程':
                        repair = list(sheet.columns)[i]
            for j in repair:
                if j.value == '累计里程':
                    repair_L.append(j)
                else:
                    if int(j.value) <= 1000000:
                        repair_L.append(j)

            repair_L1 = []
            for k in range(0, len(repair_L)):
                repair_L1.append(repair_L[k].row)

            repair_A = ''
            repair_L2 = []
            for i in range(0, len(list(sheet.columns))):
                for column in list(sheet.columns)[i]:
                    if column.value == '故障单ID':
                        repair_A = list(sheet.columns)[i]

            for i in repair_L1:
                repair_L2.append(repair_A[i - 1].value)
            print(repair_L2)
            self.log_file_out('修程修制故障单打开成功')
        except:
            self.log_file_out('修程修制故障单打开失败')

        #
        try:
            wb1 = load_workbook(r'{}/tech_fault_number/故障占比故障单.xlsx'.format(path_dir))
            sheet1 = wb1.active

            fault_a = ''
            fault_L = []
            # A1, A2, A3这样的顺序
            for i in range(0, len(list(sheet1.columns))):
                for column in list(sheet1.columns)[i]:
                    if column.value == '故障单ID':
                        fault_a = list(sheet1.columns)[i]
            self.log_file_out('故障占比故障单打开成功')
        except:
            self.log_file_out('故障占比故障单打开失败')

        for j in fault_a:
            fault_L.append(j.value)
        if len([item for item in fault_L if not item in repair_L2]) == 0:
            print('故障占比故障单与修程修制故障单一致')
        else:
            print('故障占比故障单与修程修制故障单不一致')


        Method(driver).switch_out()
        Method(driver).contains_xpath('click', '技术变更效果评估')


        Method(driver).switch_out()
        Method(driver).switch_iframe(
            driver.find_element_by_xpath("//iframe[contains(@src,'/darams/a/mould/technicalOptimization')]"))
        time.sleep(2)

        try:
            driver.find_element_by_xpath(
            "//a[contains(text(),\'{}\')]/../../td[7]/a[1]".format('导出故障单测试1')).click()
        except:
            self.log_file_out('导出技术更改故障单失败')

        time.sleep(15)
        try:
            dir_list = os.listdir(r'{}/tech_fault_number'.format(path_dir))
            dir_list = sorted(dir_list,
                              key=lambda x: os.path.getctime(os.path.join(r'{}/tech_fault_number'.format(path_dir), x)))
            os.rename(r'{}/tech_fault_number/{}'.format(path_dir, dir_list[-1]),
                      r'{}/tech_fault_number/{}'.format(path_dir, '技术更改故障单.xlsx'))
        except:
            self.log_file_out('技术更改故障单重命名失败')

        time.sleep(3)
        try:
            wb2 = load_workbook(r'{}/tech_fault_number/技术更改故障单.xlsx'.format(path_dir))
            sheet2 = wb2.active
            tech = ''
            tech_L = []
            for i in range(0, len(list(sheet2.columns))):
                for column in list(sheet2.columns)[i]:
                    if column.value == '累计里程':
                        tech = list(sheet2.columns)[i]
            for j in tech:
                if j.value == '累计里程':
                    tech_L.append(j)
                else:
                    if int(j.value) <= 1000000:
                        tech_L.append(j)

            tech_L1 = []
            for k in range(0, len(tech_L)):
                tech_L1.append(tech_L[k].row)

            tech_A = ''
            tech_L2 = []
            for i in range(0, len(list(sheet2.columns))):
                for column in list(sheet2.columns)[i]:
                    if column.value == '故障单ID':
                        tech_A = list(sheet2.columns)[i]

            for i in tech_L1:
                tech_L2.append(tech_A[i - 1].value)

            self.log_file_out('技术更改故障单打开成功')
        except:
            self.log_file_out('技术更改故障单打开失败')

        if len([item for item in fault_L if not item in tech_L2]) == 0:
            print('故障占比故障单与技术更改故障单一致')
        else:
            print('故障占比故障单与技术更改故障单不一致')

url = 'http://192.168.1.115:8080/darams/a?login'
Tech().tech_fault_analysis(url)

