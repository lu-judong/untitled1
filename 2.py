# import pymysql
# import requests
#
# class t:
#
#
#
#     def tt(self,car):
#
#         # a = ConfigurationClass().main(car,['2017-02-03','2018-05-04'])
#
#         headers = { "Content-Type": 'application/json'}
#         b = requests.post("http://192.168.1.115:8283/lcc/login",data=json.dumps({"loginName":'ljd','password':'123'}),headers=headers)
#
#         result = b.json().get('result')
#
#
#         id = self.get_id('mmmm')
#
#
#
#
#         # js1 = requests.get("http://192.168.1.115:8283/lcc/model/ratioCost/ratioChart?modelId={}&chartCode=CONFIGURATION_REPAIR_COST".format(id[0]), headers=result).json()
#         #
#         #
#         # cost = js1.get('result')
#         # c = json.loads(cost)
#         # cost1 = c.get('构型维修费用')
#         # for i in cost1:
#         #     if a.get(i[0]) == float(i[1]):
#         #         print(i[0]+'费用一致')
#         #     else:
#         #         print(i[0]+'费用不一致')
#
#
#         b = AffiliatedCost().main(car,['2017-02-03','2018-05-04'])
#         js2 = requests.get('http://192.168.1.115:8283/lcc/model/ratioCost/ratioChart?modelId={}&chartCode=AFFILIATED_REPAIR_COST'.format(id[0]),headers=result).json()
#         cost_2 = js2.get('result')
#         d = json.loads(cost_2)
#         cost2 = d.get("路局维修费用")
#         for j in cost2:
#             if b.get(j[0]) == float(j[1]):
#                 print(j[0]+'费用一致')
#             else:
#                 print(j[0]+'费用不一致')
#
#
#
# car = {'E27':['2776','2886'],
#        'E28':['2216']
#        }
#
#
# t().tt(car)


from openpyxl import load_workbook
import time


# wb = load_workbook(r'{}/tech_fault_number/{}'.format(path_dir, '故障占比故障单.xlsx'))
#
# sheet = wb.active
# # for row in sheet.rows:
# #     for cell in row:
# #         print(cell.value)
# fault = ''
# fault_L = []
# # A1, A2, A3这样的顺序
# for i in range(0,len(list(sheet.columns))):
#     for column in list(sheet.columns)[i]:
#         if column.value == '故障单ID':
#             fault = list(sheet.columns)[i]
#
# for j in fault:
#     fault_L.append(j.value)
# wb2 = load_workbook(r'{}/tech_fault_number/{}'.format(path_dir, '单一模型故障单.xlsx'))
# sheet2 = wb2.active
# singel_L = []
# # A1, A2, A3这样的顺序
# for column in list(sheet2.columns)[0]:
#     singel_L.append(column.value)
#
#
# if len([item for item in fault_L if not item in singel_L]) == 0:
#     print('单一模型与故障占比故障单数量一致')
# else:
#     print('单一模型与故障占比故障单数量不一致')

# try:
#     wb1 = load_workbook(r'{}/tech_fault_number/修程修制故障单.xlsx'.format(path_dir))
#     sheet = wb1.active
#     repair = ''
#     repair_L  = []
#     for i in range(0, len(list(sheet.columns))):
#         for column in list(sheet.columns)[i]:
#             if column.value == '累计里程':
#                 repair = list(sheet.columns)[i]
#     for j in repair:
#         if j.value == '累计里程':
#             repair_L.append(j)
#         else:
#             if int(j.value) <= 1000000:
#                 repair_L.append(j)
#
# except:
#     print('1')
# try:
#     repair_L1 = []
#     for k in range(0,len(repair_L)):
#         repair_L1.append(repair_L[k].row)
#
#     repair_A = ''
#     repair_L2 = []
#     for i in range(0, len(list(sheet.columns))):
#         for column in list(sheet.columns)[i]:
#             if column.value == '故障单ID':
#                 repair_A = list(sheet.columns)[i]
#
#
#     for i in repair_L1:
#         repair_L2.append(repair_A[i-1].value)
#     print(repair_L2)
#
# except:
#     print(2)
#
#
# wb1 = load_workbook(r'{}/tech_fault_number/故障占比故障单.xlsx'.format(path_dir))
# sheet = wb1.active
#
# fault_a = ''
# fault_L = []
# # A1, A2, A3这样的顺序
# for i in range(0, len(list(sheet.columns))):
#     for column in list(sheet.columns)[i]:
#         if column.value == '故障单ID':
#             fault_a = list(sheet.columns)[i]
#
# for j in fault_a:
#     fault_L.append(j.value)
# if len([item for item in fault_L if not item in repair_L2]) == 0:
#     print('故障占比故障单与修程修制故障单一致')
# else:
#     print('故障占比故障单与修程修制故障单不一致')

# count = 0
#
# def judge_dict(count,system):
#     try:
#         for key in system:
#             print(system[key])
#             count +=1
#             judge_dict(count,system[key])
#     except:
#         print(count)
#
# judge_dict(count,{'1':{'2':{'3':'4'}}})
import threading
import multiprocessing
from concurrent.futures import ThreadPoolExecutor
from multiprocessing import Process


pool1 = ThreadPoolExecutor(max_workers=4)

def run(b1):
    print(b1)
    time.sleep(1)



def a(len):
    threads = []

    for i in range(4):
        t = threading.Thread(target=run,args=(len[i],))
        t.start()
        threads.append(t)

    for t in threads:
        t.join()
    # print([x.result() for x in t])

    print(len,'00000000000')
def e(len):
    threads = []

    for i in range(4):
        t = threading.Thread(target=run,args=(len[i],))
        t.start()
        threads.append(t)

    for t in threads:
        t.join()
    # print([x.result() for x in t])

    print(len,'00000000000')
def c(len):
    threads = []

    for i in range(4):
        t = threading.Thread(target=run,args=(len[i],))
        t.start()
        threads.append(t)

    for t in threads:
        t.join()
    # print([x.result() for x in t])

    print(len,'00000000000')
def d(len):
    threads = []

    for i in range(4):
        t = threading.Thread(target=run,args=(len[i],))
        t.start()
        threads.append(t)

    for t in threads:
        t.join()
    # print([x.result() for x in t])

    print(len,'00000000000')

def b():
    # pool = multiprocessing.Pool(processes=4)
    # result = []
    # num = 0
    len1 = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
    # for i in range(0,4):
    #     r = pool.apply_async(a(len1[num:num+4]),args=())
    #     # p = Process(target=run,args=[])
    #     print('process {} start'.format(i))
    #     num+=4
    #     result.append(r)
    #     time.sleep(10)
    # print(result)
    # pool.close()
    p1 = Process(target=a,args=(len1[4:8],))
    p2= Process(target=e, args=(len1[0:4],))
    p3 = Process(target=c, args=(len1[8:12],))
    p4 = Process(target=d, args=(len1[12:16],))
    p1.start()
    p2.start()
    p3.start()
    p4.start()
    p1.join()
    p2.join()
    p3.join()
    p4.join()

if __name__ == '__main__':
    b()
