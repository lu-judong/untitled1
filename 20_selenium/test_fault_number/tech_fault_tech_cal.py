import os
from openpyxl import load_workbook
from config.config import *
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from config.log_config import logger
import time
from bin.main import Method
from bin.login import Login
from new_selenium.test_techm.tech_config import *
from bin.select_car import deal_car
from bin.select_fault_occur import deal_fault

#技术整改优化效果分析
class Tech:
    def log_file_out(self,msg):
        fo = open(r'{}/usecase.txt'.format(path_dir), mode='a', encoding='utf-8')
        fo.write(msg + '\r\n')
        fo.close()

    def tech_analysis(self,url,username,password,value,start,end,start1,end1,car,fault,time_sleep):
        # option.add_argument("headless")
        options = webdriver.ChromeOptions()
        prefs = {'profile.default_content_settings.popups': 0,
                 'download.default_directory': '{}\\tech_fault_number'.format(path_dir1),

                 "profile.default_content_setting_values.automatic_downloads": 1}
        options.add_experimental_option('prefs', prefs)
        driver = webdriver.Chrome(chrome_options=options,
                                  executable_path=r'{}/apps/chromedriver.exe'.format(path_dir))
        Login().login(url, username, password, driver)

        self.log_file_out('-----技术整改立即计算故障单对比-----')
        for i in contents:
            try:
                Method(driver).contains_xpath('click',i)
                time.sleep(1)
                self.log_file_out('点击'+i+'成功')
            except Exception as e:
                logger.debug(e)
                self.log_file_out('点击' + i + '失败')

        time.sleep(time_sleep)
        try:
            Method(driver).switch_out()
            Method(driver).switch_iframe(
                driver.find_element_by_xpath("//iframe[contains(@src,'/darams/a/mould/technicalOptimization')]"))
            Method(driver).click('id', 'calculate')
            time.sleep(time_sleep)
        except NoSuchElementException as e:
            self.log_file_out('点击立即计算按钮失败,获取不到相应的xpath')

        Method(driver).switch_out()
        a = Method(driver).get_attr('css', "[class='layui-layer layui-layer-iframe']", 'times')
        Method(driver).switch_iframe('layui-layer-iframe' + a)

        try:
            Method(driver).select_down_list('id', 'modelObject', value)
            Method(driver).input('id', 'confMileageList0_startMileage', start)
            Method(driver).input('id', 'confMileageList0_endMileage', end)
            Method(driver).input('id', 'confMileageList1_startMileage', start1)
            Method(driver).input('id', 'confMileageList1_endMileage', end1)
        except NoSuchElementException as e:
            print('模型录入错误 找不到对应的xpath')
        except:
            print('模型录入数据出错')

        try:
            Method(driver).contains_xpath('click', '新增')
        except NoSuchElementException as e:
            print('点击新建按钮失败')

        try:
            Method(driver).switch_out()
            b = Method(driver).get_attr('css', "[class='layui-layer layui-layer-iframe my-skin']", 'times')
            Method(driver).switch_iframe('layui-layer-iframe' + b)
        except:
            print('请录入评估对象')
            return

        time.sleep(time_sleep)

        Method(driver).click('id','partType')
        # 车型 车号新增页面
        a = deal_car(driver, car)
        if a is True:
            self.log_file_out('选车成功')
        else:
            self.log_file_out('选车失败')

        # 故障模式选择页面
        Method(driver).switch_out()
        c = Method(driver).get_attr('css', "[class='layui-layer layui-layer-iframe my-skin']", 'times')
        Method(driver).switch_iframe('layui-layer-iframe' + c)
        time.sleep(8)
        fault_status = deal_fault(driver, fault)
        if fault_status is True:
            self.log_file_out('故障模式选择成功')
        else:
            self.log_file_out('故障模式选择失败')

        try:
            Method(driver).switch_out()
            Method(driver).click('class', 'layui-layer-btn2')
            self.log_file_out('点击导出故障单按钮成功')
        except NoSuchElementException as e:
            logger.error('xpath' + '不存在!')
            self.log_file_out('点击导出故障单按钮失败')
        except:
            print('请录入完整的模型')
            return

        time.sleep(time_sleep)
        try:
            Method(driver).switch_out()
            Method(driver).switch_iframe(
                driver.find_element_by_xpath(
                    "//iframe[contains(@src,'/darams/a/mould/disposableModel/technicalOptimizationForm')]"))
            time.sleep(2)
            Method(driver).click('id', 'export')
            print('技术整改导出故障单成功')
        except NoSuchElementException:
            print('技术整改导出故障单失败')


        time.sleep(25)

        try:
            wb = load_workbook(r'{}/tech_fault_number/技术更改故障单.xlsx'.format(path_dir))
            sheet = wb.active

            tech_a = ''
            tech_L = []
            # A1, A2, A3这样的顺序
            for i in range(0, len(list(sheet.columns))):
                for column in list(sheet.columns)[i]:
                    if column.value == '故障单ID':
                        tech_a = list(sheet.columns)[i]

            for j in tech_a:
                tech_L.append(j.value)
            self.log_file_out('技术更改故障单打开成功')
        except:
            self.log_file_out('技术更改故障单打开失败')

        try:
            dir_list = os.listdir(r'{}/tech_fault_number'.format(path_dir))
            dir_list = sorted(dir_list,
                              key=lambda x: os.path.getctime(os.path.join(r'{}/tech_fault_number'.format(path_dir), x)))
            os.rename(r'{}/tech_fault_number/{}'.format(path_dir, dir_list[-1]),
                      r'{}/tech_fault_number/{}'.format(path_dir, '技术更改立即计算故障单.xlsx'))
            self.log_file_out('技术更改立即计算故障单重命名成功')
        except:
            self.log_file_out('技术更改立即计算故障单重命名失败')

        try:
            wb2 = load_workbook(r'{}/tech_fault_number/技术更改立即计算故障单.xlsx'.format(path_dir))
            sheet2 = wb2.active
            tech_cal = ''
            tech_cal_L = []
            # A1, A2, A3这样的顺序
            for i1 in range(0, len(list(sheet2.columns))):
                for column in list(sheet2.columns)[i1]:
                    if column.value == '故障单ID':
                        tech_cal = list(sheet2.columns)[i1]

            for j1 in tech_cal:
                tech_cal_L.append(j1.value)
            self.log_file_out('技术更改立即计算故障单打开成功')
        except:
            self.log_file_out('技术更改立即计算故障单打开失败')

        if len([item for item in tech_L if not item in tech_cal_L]) == 0:
            self.log_file_out('技术更改立即计算与技术更改故障单一致')
        else:
            self.log_file_out('技术更改立即计算与技术更改故障单不一致')

car = {'E27':['2641', '2642', '2643', '2644', '2645', '2646', '2647']}
fault = {'高压供电系统': 'all'}
url = 'http://192.168.1.20:8083/darams/a?login'



time_sleep = 3

Tech().tech_analysis(url,'test','1234',3,'0','100','100','200',car,fault,time_sleep)