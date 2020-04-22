from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from config.log_config import logger
import time
from bin.main import Method
from bin.login import Login
from test_fault.fault_config import *
from config.config import *
import os
from bin.select_fault_occur import deal_fault
from openpyxl import load_workbook



class Tech:
    def log_file_out(self,msg):
        fo = open(r'{}/usecase.txt'.format(path_dir), mode='a', encoding='utf-8')
        fo.write(msg + '\r\n')
        fo.close()
    
    def tech_fault_analysis(self,url,car,fault,select_fault):

        # option = webdriver.ChromeOptions()
        # option.add_argument("headless")
        # driver = webdriver.Chrome(chrome_options=option,
        #                           executable_path=r'C:\Users\a\PycharmProjects\untitled\new_selenium\apps\chromedriver.exe')

        options = webdriver.ChromeOptions()
        prefs = {'profile.default_content_settings.popups': 0,
                 'download.default_directory': '{}\\tech_fault_number'.format(path_dir1),

                 "profile.default_content_setting_values.automatic_downloads": 1}
        options.add_experimental_option('prefs', prefs)
        driver = webdriver.Chrome(chrome_options=options,
                                  executable_path=r'{}/apps/chromedriver.exe'.format(path_dir))
        Login().login(url,'test','1234',driver)

        time.sleep(2)
        for i in contents:
            try:
                Method(driver).contains_xpath('click',i)
                time.sleep(1)
                self.log_file_out('点击'+i+'成功')
            except Exception as e:
                logger.debug(e)
                self.log_file_out('点击' + i + '失败')

        time.sleep(2)

        Method(driver).switch_out()
        Method(driver).switch_iframe(
            driver.find_element_by_xpath("//iframe[contains(@src,'/darams/a/analysisProportion')]"))
        time.sleep(2)

        try:
            driver.find_element_by_xpath(
                "//a[contains(text(),\'{}\')]/../../td[7]/a[1]".format('故障占比故障单')).click()
        except:
            self.log_file_out('导出故障占比故障单失败')

        time.sleep(25)
        try:
            dir_list = os.listdir(r'{}/tech_fault_number'.format(path_dir))
            dir_list = sorted(dir_list,
                              key=lambda x: os.path.getctime(os.path.join(r'{}/tech_fault_number'.format(path_dir), x)))
            os.rename(r'{}/tech_fault_number/{}'.format(path_dir, dir_list[-1]),
                      r'{}/tech_fault_number/{}'.format(path_dir, '故障占比故障单.xlsx'))
        except:
            self.log_file_out('故障占比导出故障单重命名失败')

        time.sleep(5)
        try:
            wb = load_workbook(r'{}/tech_fault_number/故障占比故障单.xlsx'.format(path_dir))
            sheet = wb.active
            # for row in sheet.rows:
            #     for cell in row:
            #         print(cell.value)
            fault_a = ''
            fault_L = []
            # A1, A2, A3这样的顺序
            for i in range(0, len(list(sheet.columns))):
                for column in list(sheet.columns)[i]:
                    if column.value == '故障单ID':
                        fault_a = list(sheet.columns)[i]

            for j in fault_a:
                fault_L.append(j.value)
        except:
            self.log_file_out('故障占比故障单打开失败')
        time.sleep(3)
        Method(driver).switch_out()
        Method(driver).contains_xpath('click', '单一模型指标分析')

        Method(driver).switch_out()
        Method(driver).switch_iframe(
            driver.find_element_by_xpath("//iframe[contains(@src,'/darams/a/mould/singleModel')]"))
        time.sleep(2)

        try:
            driver.find_element_by_xpath(
                "//a[contains(text(),\'{}\')]/../../td[8]/a[1]".format('单一模型导出故障单')).click()
        except:
            self.log_file_out('导出单一模型故障单失败')

        time.sleep(15)

        try:
            dir_list = os.listdir(r'{}/tech_fault_number'.format(path_dir))
            dir_list = sorted(dir_list,
                              key=lambda x: os.path.getctime(os.path.join(r'{}/tech_fault_number'.format(path_dir), x)))
            os.rename(r'{}/tech_fault_number/{}'.format(path_dir, dir_list[-1]),
                      r'{}/tech_fault_number/{}'.format(path_dir, '单一模型故障单.xlsx'))
        except:
            self.log_file_out('单一模型导出故障单重命名失败')

        time.sleep(3)

        try:
            wb2 = load_workbook(r'{}/tech_fault_number/单一模型故障单.xlsx'.format(path_dir))
            sheet2 = wb2.active
            singel = ''
            singel_L = []
            # A1, A2, A3这样的顺序
            for i1 in range(0, len(list(sheet2.columns))):
                for column in list(sheet2.columns)[i1]:
                    if column.value == '故障单ID':
                        singel = list(sheet2.columns)[i1]

            for j1 in singel:
                singel_L.append(j1.value)
        except:
            self.log_file_out('单一模型故障单打开失败')


        if len([item for item in fault_L if not item in singel_L]) == 0:
            self.log_file_out('单一模型与故障占比故障单数量一致')
        else:
            self.log_file_out('单一模型与故障占比故障单数量不一致')

        Method(driver).switch_out()
        Method(driver).contains_xpath('click', '不同平台对比')

        Method(driver).switch_out()
        Method(driver).switch_iframe(
            driver.find_element_by_xpath("//iframe[contains(@src,'/darams/a/mould/moreModel')]"))
        time.sleep(2)

        try:
            driver.find_element_by_xpath(
                "//a[contains(text(),\'{}\')]/../../td[8]/a[1]".format('不同平台导出故障单')).click()
        except:
            self.log_file_out('导出不同平台故障单失败')

        time.sleep(15)
        try:
            dir_list = os.listdir(r'{}/tech_fault_number'.format(path_dir))
            dir_list = sorted(dir_list,
                              key=lambda x: os.path.getctime(os.path.join(r'{}/tech_fault_number'.format(path_dir), x)))
            os.rename(r'{}/tech_fault_number/{}'.format(path_dir, dir_list[-2]),
                      r'{}/tech_fault_number/{}'.format(path_dir, '不同平台故障单1.xlsx'))
        except:
            self.log_file_out('不同模型导出故障单重命名失败')

        time.sleep(3)
        # try:
        #     wb3 = openpyxl.load_workbook(r'{}/tech_fault_number/不同平台故障单1.xlsx'.format(path_dir))
        #     sheet3 = wb3.get_sheet_by_name('Export')
        #     more_rows1 = sheet3.max_row
        # except:
        #     self.log_file_out('打开不同平台故障单失败')
        try:
            wb3 = load_workbook(r'{}/tech_fault_number/不同平台故障单1.xlsx'.format(path_dir))
            sheet3 = wb3.active
            compare0 = ''
            compare_L1 = []
            # A1, A2, A3这样的顺序
            for i2 in range(0, len(list(sheet3.columns))):
                for column in list(sheet3.columns)[i2]:
                    if column.value == '故障单ID':
                        compare0 = list(sheet3.columns)[i2]

            for j2 in compare0:
                compare_L1.append(j2.value)
            self.log_file_out('不同平台故障单打开成功')
        except:
            self.log_file_out('不同平台故障单打开失败')

        if len([item for item in fault_L if not item in compare_L1]) == 0:
            self.log_file_out('不同平台故障单1与故障占比故障单数量一致')
        else:
            self.log_file_out('不同平台故障单1与故障占比故障单数量不一致')

        try:
            dir_list = os.listdir(r'{}/tech_fault_number'.format(path_dir))
            dir_list = sorted(dir_list,
                              key=lambda x: os.path.getctime(os.path.join(r'{}/tech_fault_number'.format(path_dir), x)))
            os.rename(r'{}/tech_fault_number/{}'.format(path_dir, dir_list[-1]),
                      r'{}/tech_fault_number/{}'.format(path_dir, '不同平台故障单2.xlsx'))
        except:
            self.log_file_out('不同模型导出故障单重命名失败')

        time.sleep(3)
        try:
            wb4 = load_workbook(r'{}/tech_fault_number/不同平台故障单2.xlsx'.format(path_dir))
            sheet4 = wb4.active
            compare1 = ''
            compare_L2 = []
            # A1, A2, A3这样的顺序
            for i3 in range(0, len(list(sheet4.columns))):
                for column in list(sheet4.columns)[i3]:
                    if column.value == '故障单ID':
                        compare1 = list(sheet4.columns)[i3]

            for j3 in compare1:
                compare_L2.append(j3.value)
            self.log_file_out('不同平台故障单打开成功')
        except:
            self.log_file_out('不同平台故障单打开失败')

        if len([item for item in fault_L if not item in compare_L2]) == 0:
            self.log_file_out('不同平台故障单2与故障占比故障单数量一致')
        else:
            self.log_file_out('不同平台故障单2与故障占比故障单数量不一致')

        Method(driver).switch_out()
        Method(driver).contains_xpath('click', 'RAMS指标评估')

        time.sleep(2)
        try:
            Method(driver).switch_out()
            Method(driver).switch_iframe(
                driver.find_element_by_xpath("//iframe[contains(@src,'/darams/a/mould/disposableModel/form')]"))
            time.sleep(2)
        except NoSuchElementException as e:
            self.log_file_out('切入iframe失败')
        try:
            Method(driver).select_down_list('id', 'modelObject', 3)
            Method(driver).input('id', 'confMileageList0_startMileage', '0')
            Method(driver).input('id', 'confMileageList0_endMileage', '100')
        except NoSuchElementException as e:
            self.log_file_out('模型基本信息输入失败')
        except:
            self.log_file_out('请录入评估对象')
            return

        try:
            Method(driver).contains_xpath('click','新增')
        except NoSuchElementException as e:
            self.log_file_out('点击新建按钮失败')

        try:
            Method(driver).switch_out()
            b = Method(driver).get_attr('css',"[class='layui-layer layui-layer-iframe my-skin']", 'times')
            Method(driver).switch_iframe('layui-layer-iframe' + b)
        except:
            self.log_file_out('请录入评估对象')
            return

        time.sleep(3)

        # 车型 车号新增页面
        if select_fault == '交集':
            js_s = "return $('#partType').prop('checked')"
            status = driver.execute_script(js_s)
            if status is True:
                Method(driver).click('id', 'partType')
            else:
                time.sleep(1)
        elif select_fault == '并集':
            js_s = "return $('#partType').prop('checked')"
            status = driver.execute_script(js_s)
            if status is True:
                time.sleep(1)
            else:
                Method(driver).click('id', 'partType')

        try:
            for x in car:
                driver.find_element_by_xpath(
                    "//span[@class='train zoomIn' and contains(text(),\'{}\')]".format(x)).click()
                car_num = car.get(x)
                time.sleep(2)
                if car_num == 'all':
                    driver.execute_script('top.select_all(this,"train-no-layer-class",true,true)')
                else:
                    for i in car_num:
                        # driver.find_element_by_xpath("//span[contains(text(),\'{}\')]".format(i)).click()
                        car_num = driver.find_element_by_xpath(
                            '//span[contains(text(),\'{}\')]/..'.format(i)).get_attribute('id')

                        js = 'top.checked(top.$("#{}"))'.format(car_num)
                        driver.execute_script(js)

                    # time.sleep(2)
                driver.execute_script('top.$(".layui-layer-btn0")[1].click()')
                time.sleep(2)
            Method(driver).switch_out()

            Method(driver).two_element_click("[class='layui-layer layui-layer-iframe my-skin']", 'layui-layer-btn1')
        except NoSuchElementException as e:
            logger.error('xpath' + '不存在!')


        # 故障模式选择页面
        Method(driver).switch_out()
        c = Method(driver).get_attr('css',"[class='layui-layer layui-layer-iframe my-skin']", 'times')
        Method(driver).switch_iframe('layui-layer-iframe' + c)
        time.sleep(5)
        fault_status = deal_fault(driver,fault)
        if fault_status is True:
            self.log_file_out('故障模式选择成功')
        else:
            self.log_file_out('故障模式选择失败')

        try:
            Method(driver).switch_out()
            Method(driver).click('class', 'layui-layer-btn2')
        except NoSuchElementException as e:
            logger.error('xpath' + '不存在!')
            self.log_file_out('点击确定按钮失败')
        except:
            self.log_file_out('请录入完整的模型')
            return


        time.sleep(2)
        try:
            Method(driver).switch_out()
            Method(driver).switch_iframe(
                driver.find_element_by_xpath("//iframe[contains(@src,'/darams/a/mould/disposableModel/form')]"))
            Method(driver).click('id', 'export')
            self.log_file_out('点击导出故障单按钮成功')
        except:
            self.log_file_out('点击导出故障单按钮失败')

        time.sleep(15)

        try:
            dir_list = os.listdir(r'{}/tech_fault_number'.format(path_dir))
            dir_list = sorted(dir_list,
                              key=lambda x: os.path.getctime(os.path.join(r'{}/tech_fault_number'.format(path_dir), x)))
            os.rename(r'{}/tech_fault_number/{}'.format(path_dir, dir_list[-1]),
                      r'{}/tech_fault_number/{}'.format(path_dir, 'RAMS故障单.xlsx'))
        except:
            self.log_file_out('RAMS导出故障单重命名失败')

        time.sleep(3)
        try:
            wb5 = load_workbook(r'{}/tech_fault_number/RAMS故障单.xlsx'.format(path_dir))
            sheet5 = wb5.active
            RAMS = ''
            RAMS_L = []
            # A1, A2, A3这样的顺序
            for i4 in range(0, len(list(sheet5.columns))):
                for column in list(sheet5.columns)[i4]:
                    if column.value == '故障单ID':
                        RAMS = list(sheet5.columns)[i4]

            for j1 in RAMS:
                RAMS_L.append(j1.value)
            self.log_file_out('RAMS故障单打开成功')
        except:
            self.log_file_out('RAMS故障单打开失败')

        if len([item for item in fault_L if not item in RAMS_L]) == 0:
            self.log_file_out('RAMS二合一故障单与故障占比故障单数量一致')
        else:
            self.log_file_out('RMAS二合一故障单与故障占比故障单数量不一致')

car = {'E27':['2641', '2642', '2643', '2644', '2645', '2646', '2647']}
fault = {'高压供电系统': 'all'}
url = 'http://192.168.1.20:8083/darams/a?login'
Tech().tech_fault_analysis(url,car,fault,'交集')