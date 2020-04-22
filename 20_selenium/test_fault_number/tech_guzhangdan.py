import os
from openpyxl import load_workbook
from config.config import *
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from config.log_config import logger
import time
from bin.main import Method
from bin.login import Login

from bin.select_car import deal_car


#技术整改优化效果分析
class Tech:
    def log_file_out(self,msg):
        fo = open(r'{}/usecase.txt'.format(path_dir), mode='a', encoding='utf-8')
        fo.write(msg + '\r\n')
        fo.close()

    def tech_analysis(self,url,username,password,car):

        options = webdriver.ChromeOptions()
        prefs = {'profile.default_content_settings.popups': 0,
                 'download.default_directory': '{}\\tech_fault_number'.format(path_dir1),
                 "profile.default_content_setting_values.automatic_downloads": 1}
        options.add_experimental_option('prefs', prefs)
        driver = webdriver.Chrome(chrome_options=options,
                                  executable_path=r'{}/apps/chromedriver.exe'.format(path_dir))
        Login().login(url, username, password, driver)

        self.log_file_out('-----技术整改立即计算故障单对比-----')
        try:
            Method(driver).contains_xpath('click', '基础数据')
            time.sleep(1)
            Method(driver).contains_xpath('click', '故障单')
            self.log_file_out('点击故障单菜单成功')
        except:
            self.log_file_out('点击故障单菜单失败')

        Method(driver).switch_out()
        Method(driver).switch_iframe(
            driver.find_element_by_xpath("//iframe[contains(@src,'/darams/a/fault/opFaultOrder')]"))

        time.sleep(1)
        try:
            Method(driver).click('xpath','//*[@id="toolbar"]/a[2]')
            time.sleep(3)
            Method(driver).input('id','startMileage',0)
            time.sleep(1)
            Method(driver).input('id', 'endMileage', 100)
            self.log_file_out('基本信息录入成功')
        except:
            self.log_file_out('基本信息录入失败')
        time.sleep(1)
        Method(driver).click('id','trainSelectButton')

        Method(driver).switch_out()
        Method(driver).switch_iframe(
            driver.find_element_by_xpath("//iframe[contains(@src,'/darams/a/mould/confModel/trainGroup')]"))
        try:
            for x in car:
                driver.find_element_by_xpath(
                    "//span[@class='train zoomIn' and contains(text(),\'{}\')]".format(x)).click()
                car_num = car.get(x)
                time.sleep(2)
                for i in car_num:
                    # driver.find_element_by_xpath("//span[contains(text(),\'{}\')]".format(i)).click()
                    car_num = driver.find_element_by_xpath('//span[contains(text(),\'{}\')]/..'.format(i)).get_attribute(
                        'id')
                    js = 'top.checked(top.$("#{}"))'.format(car_num)
                    driver.execute_script(js)
                driver.execute_script('top.$(".layui-layer-btn0")[1].click()')
                time.sleep(2)
                driver.execute_script('top.$(".layui-layer-btn0")[0].click()')
        except NoSuchElementException as e:
            logger.error('xpath' + '不存在!')

        Method(driver).switch_out()
        Method(driver).switch_iframe(
            driver.find_element_by_xpath("//iframe[contains(@src,'/darams/a/fault/opFaultOrder')]"))
        Method(driver).click('id','systemPartSelectButton')

        Method(driver).switch_out()
        Method(driver).switch_iframe(
            driver.find_element_by_xpath("//iframe[contains(@src,'/darams/a/basic/cdFaultObject/fault_obj_view')]"))
        time.sleep(1)
        Method(driver).click('xpath','//*[@id="841A2D86F7C3B6C9E050A8C00C01289E000_anchor"]')
        time.sleep(1)
        Method(driver).click('xpath','//*[@id="middleTable"]/tbody/tr[1]/td[1]/input')
        time.sleep(1)
        Method(driver).click('xpath', '//*[@id="middleTable"]/tbody/tr[2]/td[1]/input')
        time.sleep(1)
        Method(driver).click('xpath', '//*[@id="middleTable"]/tbody/tr[7]/td[1]/input')
        time.sleep(1)
        Method(driver).click('xpath', '//*[@id="middleTable"]/tbody/tr[8]/td[1]/input')
        time.sleep(1)
        Method(driver).click('xpath', '//*[@id="middleTable"]/tbody/tr[10]/td[1]/input')
        time.sleep(1)
        Method(driver).click('xpath','//*[@id="fault-body"]/table/tbody/tr/td[3]/div[1]/div[2]/div[4]/div[2]/ul/li[4]/a')
        time.sleep(2)
        Method(driver).click('xpath', '//*[@id="middleTable"]/tbody/tr[1]/td[1]/input')
        time.sleep(1)
        Method(driver).click('xpath', '//*[@id="middleTable"]/tbody/tr[2]/td[1]/input')
        time.sleep(1)
        Method(driver).click('xpath', '//*[@id="middleTable"]/tbody/tr[4]/td[1]/input')
        time.sleep(1)
        Method(driver).click('xpath', '//*[@id="middleTable"]/tbody/tr[5]/td[1]/input')
        time.sleep(1)
        Method(driver).click('xpath', '//*[@id="middleTable"]/tbody/tr[6]/td[1]/input')
        time.sleep(1)
        Method(driver).click('id','right-move')
        time.sleep(1)
        Method(driver).click('id','isContainChildComponent')
        driver.execute_script('top.$(".layui-layer-btn0")[0].click()')
        Method(driver).switch_out()
        Method(driver).switch_iframe(
            driver.find_element_by_xpath("//iframe[contains(@src,'/darams/a/fault/opFaultOrder')]"))
        time.sleep(1)
        Method(driver).click('id','export')
        Method(driver).switch_out()

        driver.find_element_by_class_name('layui-layer-btn0').click()
        time.sleep(10)

        try:
            wb = load_workbook(r'{}/tech_fault_number/故障占比故障单.xlsx'.format(path_dir))
            sheet = wb.active
            # for row in sheet.rows:
            #     for cell in row:
            #         print(cell.value)
            fault_a = ''
            fault_L = []
            # A1, A2, A3这样的顺序
            for i in range(0, len(list(sheet.columns))):
                for column in list(sheet.columns)[i]:
                    if column.value == '故障单ID':
                        fault_a = list(sheet.columns)[i]

            for j in fault_a:
                fault_L.append(j.value)
        except:
            self.log_file_out('故障占比故障单打开失败')

        try:
            dir_list = os.listdir(r'{}/tech_fault_number'.format(path_dir))
            dir_list = sorted(dir_list,
                              key=lambda x: os.path.getctime(os.path.join(r'{}/tech_fault_number'.format(path_dir), x)))
            os.rename(r'{}/tech_fault_number/{}'.format(path_dir, dir_list[-1]),
                      r'{}/tech_fault_number/{}'.format(path_dir, '故障单功能的故障单.xlsx'))
        except:
            self.log_file_out('故障单功能导出故障单重命名失败')

        time.sleep(3)

        try:
            wb2 = load_workbook(r'{}/tech_fault_number/故障单功能的故障单.xlsx'.format(path_dir))
            sheet2 = wb2.active
            faultl = ''
            fault_L1 = []
            # A1, A2, A3这样的顺序
            for i1 in range(0, len(list(sheet2.columns))):
                for column in list(sheet2.columns)[i1]:
                    if column.value == '故障单ID':
                        faultl = list(sheet2.columns)[i1]

            for j1 in faultl:
                fault_L1.append(j1.value)
        except:
            self.log_file_out('故障单功能故障单打开失败')

        if len([item for item in fault_L if not item in fault_L1]) == 0:
            self.log_file_out('故障单功能与故障占比故障单数量一致')
        else:
            self.log_file_out('故障单功能与故障占比故障单数量不一致')




car = {'E27':['2641', '2642', '2643', '2644', '2645', '2646', '2647']}

url = 'http://192.168.1.115:8080/darams/a?login'



time_sleep = 3

Tech().tech_analysis(url,'test','1234',car)

